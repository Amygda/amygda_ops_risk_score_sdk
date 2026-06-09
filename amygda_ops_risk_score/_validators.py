"""Client-side validators — all checks run before any HTTP call.

Every public function raises ``ValidationError`` on failure.
The API is never called unless all checks pass.
"""

from __future__ import annotations

import os
import zipfile
from typing import List, Optional

from amygda_ops_risk_score.exceptions import ValidationError


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _raise(errors: list) -> None:
    raise ValidationError(errors)


def _xlsx_header(file_bytes: bytes, sheet_name: Optional[str]) -> list:
    """Return column headers from an XLSX/XLS file.  Raises ValidationError on sheet issues."""
    try:
        import openpyxl
    except ImportError:
        return []

    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if sheet_name is None:
        if len(sheet_names) > 1:
            available = ", ".join(f'"{s}"' for s in sheet_names)
            _raise([
                f"This workbook has multiple sheets: {available}. "
                "Specify sheet_name to choose one."
            ])
        ws = wb.worksheets[0]
    else:
        if sheet_name not in sheet_names:
            available = ", ".join(f'"{s}"' for s in sheet_names)
            _raise([
                f"Sheet '{sheet_name}' not found in this workbook. "
                f"Available sheets: {available}"
            ])
        ws = wb[sheet_name]

    rows = list(ws.iter_rows(max_row=1, values_only=True))
    if not rows:
        return []
    return [str(c) for c in rows[0] if c is not None]


# ---------------------------------------------------------------------------
# configure (Step 1)
# ---------------------------------------------------------------------------

def validate_configure(
    file_path: str,
    log_column: str,
    max_systems: int,
    max_subsystems: int,
    asset_context: str,
    is_free_text: bool,
    sheet_name: Optional[str] = None,
) -> None:
    errors: list = []

    # File checks
    if not os.path.exists(file_path):
        _raise([f"File not found: {file_path}"])

    size_mb = os.path.getsize(file_path) / (1024 * 1024)
    if size_mb > 100:
        errors.append(f"File too large ({size_mb:.1f} MB). Maximum is 100 MB.")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in {".csv", ".xlsx", ".xls"}:
        errors.append(f"Unsupported file type '{ext}'. Must be .csv, .xlsx, or .xls.")

    if errors:
        _raise(errors)

    # Parameter checks
    if not log_column or not log_column.strip():
        errors.append("log_column is required and cannot be blank.")

    if not asset_context or not asset_context.strip():
        errors.append("asset_context is required and cannot be blank.")
    else:
        # Prompt injection guard — asset_context is used verbatim in LLM prompts
        _INJECTION_PATTERNS = [
            "ignore previous", "ignore all", "disregard", "forget everything",
            "system prompt", "jailbreak", "you are now", "act as",
        ]
        lower_ctx = asset_context.lower()
        for pattern in _INJECTION_PATTERNS:
            if pattern in lower_ctx:
                errors.append(
                    f"asset_context contains a potentially unsafe phrase: '{pattern}'. "
                    "Provide a plain description of the asset context."
                )
                break

    if not isinstance(is_free_text, bool):
        errors.append("is_free_text must be True or False.")

    if not isinstance(max_systems, int) or not (1 <= max_systems <= 10):
        errors.append("max_systems must be an integer between 1 and 10.")

    if not isinstance(max_subsystems, int) or not (1 <= max_subsystems <= 5):
        errors.append("max_subsystems must be an integer between 1 and 5.")

    if errors:
        _raise(errors)

    # Best-effort column check (skip if openpyxl absent for CSV)
    if ext in {".xlsx", ".xls"}:
        with open(file_path, "rb") as fh:
            file_bytes = fh.read()
        headers = _xlsx_header(file_bytes, sheet_name)
        if headers and log_column not in headers:
            errors.append(
                f"Column '{log_column}' not found in file headers. "
                f"Available columns: {', '.join(headers)}"
            )
    elif ext == ".csv":
        try:
            import pandas as pd
            df = pd.read_csv(file_path, nrows=0)
            if log_column not in df.columns:
                errors.append(
                    f"Column '{log_column}' not found in CSV headers. "
                    f"Available columns: {', '.join(df.columns)}"
                )
        except Exception:
            pass  # best-effort only

    if errors:
        _raise(errors)


# ---------------------------------------------------------------------------
# downsample (Step 2)
# ---------------------------------------------------------------------------

def validate_downsample(
    sample_size: int,
    asset_column: Optional[str] = None,
    vehicle_column: Optional[str] = None,
    timestamp_column: Optional[str] = None,
    file_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> None:
    errors: list = []

    if not isinstance(sample_size, int) or sample_size < 1:
        errors.append("sample_size must be a positive integer.")
    elif sample_size > 5000:
        errors.append(f"sample_size {sample_size} exceeds maximum of 5000.")

    provided = [c for c in [asset_column, vehicle_column, timestamp_column] if c and c.strip()]
    if not provided:
        errors.append(
            "At least one stratification column is required. "
            "Provide one or more of: asset_column, vehicle_column, timestamp_column."
        )

    if errors:
        _raise(errors)

    # Best-effort column name check against the source file (if provided)
    if file_path and provided and os.path.exists(file_path):
        ext = os.path.splitext(file_path)[1].lower()
        headers: list = []

        if ext in {".xlsx", ".xls"}:
            try:
                with open(file_path, "rb") as fh:
                    headers = _xlsx_header(fh.read(), sheet_name)
            except Exception:
                pass
        elif ext == ".csv":
            try:
                import pandas as pd
                headers = list(pd.read_csv(file_path, nrows=0).columns)
            except Exception:
                pass

        if headers:
            for col in provided:
                if col not in headers:
                    errors.append(
                        f"Column '{col}' not found in dataset. "
                        f"Available columns: {', '.join(headers)}"
                    )

    if errors:
        _raise(errors)


# ---------------------------------------------------------------------------
# extract_keywords (Step 3)
# ---------------------------------------------------------------------------

def validate_extract_keywords(extraction_method: str) -> None:
    if extraction_method not in {"fast", "deep"}:
        _raise([
            f"extraction_method '{extraction_method}' is not supported. "
            "Use 'fast' (quick, recommended) or 'deep' (slower, better for noisy text)."
        ])


# ---------------------------------------------------------------------------
# generate_hierarchy (Step 4)
# ---------------------------------------------------------------------------

def validate_generate_hierarchy(keywords: Optional[List[str]] = None) -> None:
    if keywords is None:
        return
    if not isinstance(keywords, list) or len(keywords) == 0:
        _raise(["keywords must be a non-empty list of strings when provided."])
    blank = [k for k in keywords if not isinstance(k, str) or not k.strip()]
    if blank:
        _raise(["All keywords must be non-blank strings."])


# ---------------------------------------------------------------------------
# update_hierarchy (Step 5)
# ---------------------------------------------------------------------------

_CONFIDENCE_ENUMS = {"high", "medium", "low"}
_ALLOWED_HIERARCHY_FIELDS = {"system", "system_confidence", "subsystem", "subsystem_confidence"}


def validate_update_hierarchy(rows: List[dict]) -> None:
    errors: list = []

    if not rows:
        errors.append("rows must be a non-empty list.")
        _raise(errors)

    for i, row in enumerate(rows):
        # Reject unknown fields — keywords in particular is a common mistake
        unknown = set(row.keys()) - _ALLOWED_HIERARCHY_FIELDS
        if unknown:
            tip = (
                "\n    Tip: 'keywords' is set at extract_keywords and cannot be edited here."
                if "keywords" in unknown else ""
            )
            errors.append(
                f"Row {i}: unexpected field(s): {sorted(unknown)}\n"
                f"    Allowed fields: {sorted(_ALLOWED_HIERARCHY_FIELDS)}{tip}"
            )
            continue  # skip further checks on this row

        missing = _ALLOWED_HIERARCHY_FIELDS - set(row.keys())
        if missing:
            errors.append(
                f"Row {i}: missing required field(s): {sorted(missing)}\n"
                f"    Required: system, system_confidence, subsystem, subsystem_confidence"
            )
            continue

        if not row["system"] or not str(row["system"]).strip():
            errors.append(
                f"Row {i}: 'system' is blank or null.\n"
                f"    Got: {row['system']!r}  →  provide a non-empty system name."
            )
        if not row["subsystem"] or not str(row["subsystem"]).strip():
            errors.append(
                f"Row {i}: 'subsystem' is blank or null.\n"
                f"    Got: {row['subsystem']!r}  →  provide a non-empty subsystem name."
            )

        # Confidence enums are normalised to lowercase — "HIGH" / "High" / "high" all accepted.
        sc = str(row.get("system_confidence", "")).strip().lower()
        if sc not in _CONFIDENCE_ENUMS:
            errors.append(
                f"Row {i}: 'system_confidence' is invalid.\n"
                f"    Got: {row.get('system_confidence')!r}  →  must be 'high', 'medium', or 'low'."
            )

        ssc = str(row.get("subsystem_confidence", "")).strip().lower()
        if ssc not in _CONFIDENCE_ENUMS:
            errors.append(
                f"Row {i}: 'subsystem_confidence' is invalid.\n"
                f"    Got: {row.get('subsystem_confidence')!r}  →  must be 'high', 'medium', or 'low'."
            )

    if errors:
        _raise(errors)

    # Duplicate (system, subsystem) pairs produce silent data corruption — reject early.
    seen: set = set()
    for i, row in enumerate(rows):
        key = (str(row.get("system", "")).strip(), str(row.get("subsystem", "")).strip())
        if key in seen:
            errors.append(
                f"Row {i}: duplicate entry for system='{key[0]}', subsystem='{key[1]}'. "
                "Each (system, subsystem) pair must appear only once."
            )
        seen.add(key)

    if errors:
        _raise(errors)

    # system_confidence is a system-level attribute — all rows for the same system
    # must carry the same value.  subsystem_confidence is per-row and may differ freely.
    system_conf_map: dict = {}
    for i, row in enumerate(rows):
        sys  = str(row.get("system", "")).strip()
        conf = str(row.get("system_confidence", "")).strip().lower()
        if sys in system_conf_map:
            if system_conf_map[sys] != conf:
                errors.append(
                    f"Row {i}: system_confidence for '{sys}' is inconsistent — "
                    f"found '{system_conf_map[sys]}' and '{conf}'.\n"
                    f"    system_confidence is a system-level attribute: all rows for "
                    f"the same system must have the same value.\n"
                    f"    subsystem_confidence is per-row and may differ across subsystems."
                )
        else:
            system_conf_map[sys] = conf

    if errors:
        _raise(errors)


# ---------------------------------------------------------------------------
# update_weights (Step 7)
# ---------------------------------------------------------------------------

def validate_update_weights(
    systems: List[dict],
    original_systems: Optional[List[dict]] = None,
) -> None:
    errors: list = []

    if not systems:
        errors.append("systems must be a non-empty list.")
        _raise(errors)

    for i, sys_entry in enumerate(systems):
        if not sys_entry.get("system_name") or not str(sys_entry["system_name"]).strip():
            errors.append(f"System {i}: system_name cannot be blank.")

        w = sys_entry.get("weight")
        sname = sys_entry.get("system_name", f"index {i}")
        if not isinstance(w, (int, float)):
            errors.append(
                f"System '{sname}': weight must be an int or float.\n"
                f"    Got: {w!r} ({type(w).__name__})  →  e.g. 0.6"
            )
        elif not (0.01 <= w <= 1.0):
            errors.append(
                f"System '{sname}': weight out of range.\n"
                f"    Got: {w}  →  must be between 0.01 and 1.0."
            )

        subs = sys_entry.get("subsystems", [])
        if not subs:
            errors.append(
                f"System '{sname}': subsystems list is empty.\n"
                f"    Each system must have at least one subsystem."
            )
            continue

        for j, sub in enumerate(subs):
            subname = sub.get("subsystem_name", f"index {j}")
            if not sub.get("subsystem_name") or not str(sub["subsystem_name"]).strip():
                errors.append(
                    f"System '{sname}', subsystem {j}: subsystem_name is blank or null.\n"
                    f"    Got: {sub.get('subsystem_name')!r}  →  provide a non-empty name."
                )
            sw = sub.get("weight")
            if not isinstance(sw, (int, float)):
                errors.append(
                    f"System '{sname}', subsystem '{subname}': weight must be an int or float.\n"
                    f"    Got: {sw!r} ({type(sw).__name__})  →  e.g. 0.4"
                )
            elif not (0.01 <= sw <= 1.0):
                errors.append(
                    f"System '{sname}', subsystem '{subname}': weight out of range.\n"
                    f"    Got: {sw}  →  must be between 0.01 and 1.0."
                )

    if errors:
        _raise(errors)

    # Name-change lock: system/subsystem names must match the original exactly.
    # Only weight values are allowed to differ.
    if original_systems is not None:
        orig_sys_names = [s["system_name"] for s in original_systems]
        new_sys_names  = [s["system_name"] for s in systems]
        if orig_sys_names != new_sys_names:
            errors.append(
                "System names cannot be changed in update_weights — only weight values may differ. "
                f"Expected systems (in order): {orig_sys_names}. "
                f"Got: {new_sys_names}. "
                "Use helpers.make_weight_update() to get a safely editable copy."
            )
        else:
            orig_map = {s["system_name"]: s for s in original_systems}
            for sys_entry in systems:
                sname = sys_entry["system_name"]
                orig_subs = [sub["subsystem_name"] for sub in orig_map[sname].get("subsystems", [])]
                new_subs  = [sub["subsystem_name"] for sub in sys_entry.get("subsystems", [])]
                if orig_subs != new_subs:
                    errors.append(
                        f"System '{sname}': subsystem names cannot be changed — only weight values may differ. "
                        f"Expected: {orig_subs}. Got: {new_subs}."
                    )
        if errors:
            _raise(errors)

    # Sum checks — 0.0001 tolerance handles float rounding (e.g. 0.33+0.33+0.34)
    sys_total = sum(float(s.get("weight", 0)) for s in systems)
    if abs(sys_total - 1.0) > 0.0001:
        breakdown = ", ".join(
            f"'{s.get('system_name', '?')}' = {s.get('weight')}" for s in systems
        )
        errors.append(
            f"System weights sum to {sys_total:.4f}, must be exactly 1.0.\n"
            f"    Breakdown: {breakdown}"
        )

    for i, sys_entry in enumerate(systems):
        subs = sys_entry.get("subsystems", [])
        if not subs:
            continue
        sub_total = sum(float(s.get("weight", 0)) for s in subs)
        if abs(sub_total - 1.0) > 0.0001:
            breakdown = ", ".join(
                f"'{s.get('subsystem_name', '?')}' = {s.get('weight')}" for s in subs
            )
            errors.append(
                f"System '{sys_entry.get('system_name', i)}': "
                f"subsystem weights sum to {sub_total:.4f}, must be 1.0.\n"
                f"    Breakdown: {breakdown}"
            )

    if errors:
        _raise(errors)


# ---------------------------------------------------------------------------
# import_model (Risk Score — Path B)
# ---------------------------------------------------------------------------

def validate_import_model(zip_path: str) -> None:
    errors: list = []

    if not os.path.exists(zip_path):
        _raise([f"File not found: {zip_path}"])

    ext = os.path.splitext(zip_path)[1].lower()
    if ext != ".zip":
        errors.append(f"File must be a .zip archive, got '{ext}'.")
        _raise(errors)

    if not zipfile.is_zipfile(zip_path):
        _raise(["File is not a valid zip archive (corrupt or incomplete download)."])

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()

    if "outputs/model_config.json" not in names:
        errors.append(
            "Zip is missing 'outputs/model_config.json'. "
            "Use the zip downloaded via session.download_zip() — do not modify or re-package it."
        )

    if errors:
        _raise(errors)


# ---------------------------------------------------------------------------
# configure_training (Risk Score — Step 9)
# ---------------------------------------------------------------------------

def validate_configure_training(
    file_path: str,
    asset_id_column: str,
    timestamp_column: str,
    rolling_window: int = 7,
    rolling_feature_type: str = "sum",
    quantile_for_thresholds: float = 0.99,
    sheet_name: Optional[str] = None,
    # Supervised-training params (Shery's extension)
    supervised: bool = False,
    failures_path: Optional[str] = None,
    failure_date_column: str = "machine_initialisation_date",
    sampling_strategy: str = "block",
    prediction_horizon_days: int = 14,
    exclusion_days_after: int = 7,
    target_imbalance_ratio: Optional[float] = 10.0,
    block_size_days: int = 14,
    n_candidates: int = 200,
    fallback_quantile: float = 0.95,
    min_positives: int = 5,
    lr_c: float = 0.5,
) -> None:
    errors: list = []

    if not os.path.exists(file_path):
        _raise([f"File not found: {file_path}"])

    size_mb = os.path.getsize(file_path) / (1024 * 1024)
    if size_mb > 100:
        errors.append(f"File too large ({size_mb:.1f} MB). Maximum is 100 MB.")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in {".csv", ".xlsx", ".xls"}:
        errors.append(f"Unsupported file type '{ext}'. Must be .csv, .xlsx, or .xls.")

    if not asset_id_column or not asset_id_column.strip():
        errors.append("asset_id_column is required and cannot be blank.")

    if not timestamp_column or not timestamp_column.strip():
        errors.append("timestamp_column is required and cannot be blank.")

    if not isinstance(rolling_window, int) or not (1 <= rolling_window <= 30):
        errors.append("rolling_window must be an integer between 1 and 30.")

    if rolling_feature_type not in {"sum", "flag", "ewm", "all"}:
        errors.append(
            f"rolling_feature_type '{rolling_feature_type}' is invalid. Must be 'sum', 'flag', 'ewm', or 'all'."
        )

    try:
        q = float(quantile_for_thresholds)
        if not (0.5 <= q < 1.0):
            errors.append("quantile_for_thresholds must be a float in [0.5, 1.0).")
    except (TypeError, ValueError):
        errors.append("quantile_for_thresholds must be a number.")

    # Supervised-training param validation
    if not isinstance(supervised, bool):
        errors.append("supervised must be a bool (True or False).")

    if supervised:
        if not failures_path:
            errors.append(
                "failures_path is required when supervised=True. "
                "Provide a CSV path containing failure event dates."
            )
        if not failure_date_column or not failure_date_column.strip():
            errors.append("failure_date_column is required and cannot be blank.")

    if sampling_strategy not in {"block", "random"}:
        errors.append(
            f"sampling_strategy '{sampling_strategy}' is invalid. Must be 'block' or 'random'."
        )

    if not isinstance(prediction_horizon_days, int) or prediction_horizon_days <= 0:
        errors.append("prediction_horizon_days must be a positive integer.")

    if not isinstance(exclusion_days_after, int) or exclusion_days_after < 0:
        errors.append("exclusion_days_after must be a non-negative integer.")

    if target_imbalance_ratio is not None:
        try:
            r = float(target_imbalance_ratio)
            if r <= 0:
                errors.append("target_imbalance_ratio must be a positive number, or None to keep all negatives.")
        except (TypeError, ValueError):
            errors.append("target_imbalance_ratio must be a positive number, or None.")

    if not isinstance(block_size_days, int) or block_size_days <= 0:
        errors.append("block_size_days must be a positive integer.")

    if not isinstance(n_candidates, int) or n_candidates <= 0:
        errors.append("n_candidates must be a positive integer.")

    try:
        fq = float(fallback_quantile)
        if not (0.5 <= fq < 1.0):
            errors.append("fallback_quantile must be a float in [0.5, 1.0).")
    except (TypeError, ValueError):
        errors.append("fallback_quantile must be a number.")

    if not isinstance(min_positives, int) or min_positives < 1:
        errors.append("min_positives must be an integer >= 1.")

    try:
        c = float(lr_c)
        if c <= 0:
            errors.append("lr_c must be a positive number.")
    except (TypeError, ValueError):
        errors.append("lr_c must be a positive number.")

    if errors:
        _raise(errors)

    # Best-effort column presence check — main training file
    if ext in {".xlsx", ".xls"}:
        with open(file_path, "rb") as fh:
            file_bytes = fh.read()
        headers = _xlsx_header(file_bytes, sheet_name)
        if headers:
            for col in [asset_id_column, timestamp_column]:
                if col not in headers:
                    errors.append(
                        f"Column '{col}' not found in file headers. "
                        f"Available: {', '.join(headers)}"
                    )
    elif ext == ".csv":
        try:
            import pandas as pd
            df = pd.read_csv(file_path, nrows=0)
            for col in [asset_id_column, timestamp_column]:
                if col not in df.columns:
                    errors.append(
                        f"Column '{col}' not found in CSV headers. "
                        f"Available: {', '.join(df.columns)}"
                    )
        except Exception:
            pass

    # Best-effort column presence check — failures file (supervised=True only)
    if supervised and failures_path and os.path.exists(failures_path):
        fp_ext = os.path.splitext(failures_path)[1].lower()

        fp_size_mb = os.path.getsize(failures_path) / (1024 * 1024)
        if fp_size_mb > 100:
            errors.append(
                f"failures_path file too large ({fp_size_mb:.1f} MB). Maximum is 100 MB."
            )

        if fp_ext not in {".csv", ".xlsx", ".xls"}:
            errors.append(
                f"failures_path has unsupported file type '{fp_ext}'. Must be .csv, .xlsx, or .xls."
            )
        elif fp_ext == ".csv":
            try:
                import pandas as pd
                fp_df = pd.read_csv(failures_path, nrows=0)
                if failure_date_column not in fp_df.columns:
                    errors.append(
                        f"Column '{failure_date_column}' not found in failures_path CSV. "
                        f"Available: {', '.join(fp_df.columns)}"
                    )
            except Exception:
                pass
        elif fp_ext in {".xlsx", ".xls"}:
            with open(failures_path, "rb") as fh:
                fp_bytes = fh.read()
            fp_headers = _xlsx_header(fp_bytes, None)
            if fp_headers and failure_date_column not in fp_headers:
                errors.append(
                    f"Column '{failure_date_column}' not found in failures_path headers. "
                    f"Available: {', '.join(fp_headers)}"
                )
    elif supervised and failures_path and not os.path.exists(failures_path):
        errors.append(f"failures_path not found: {failures_path}")

    if errors:
        _raise(errors)


# ---------------------------------------------------------------------------
# configure_generation (Risk Score — Step 11)
# ---------------------------------------------------------------------------

def validate_configure_generation(
    file_path: str,
    sheet_name: Optional[str] = None,
    weights_source: str = "labelling",
) -> None:
    errors: list = []

    if not os.path.exists(file_path):
        _raise([f"File not found: {file_path}"])

    size_mb = os.path.getsize(file_path) / (1024 * 1024)
    if size_mb > 100:
        errors.append(f"File too large ({size_mb:.1f} MB). Maximum is 100 MB.")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in {".csv", ".xlsx", ".xls"}:
        errors.append(f"Unsupported file type '{ext}'. Must be .csv, .xlsx, or .xls.")
        _raise(errors)

    if weights_source not in {"labelling", "supervised"}:
        errors.append(
            f"weights_source '{weights_source}' is invalid. Must be 'labelling' or 'supervised'."
        )

    if errors:
        _raise(errors)

    # Validate sheet selection for XLSX
    if ext in {".xlsx", ".xls"}:
        with open(file_path, "rb") as fh:
            file_bytes = fh.read()
        _xlsx_header(file_bytes, sheet_name)  # raises if sheet issues

    if errors:
        _raise(errors)
